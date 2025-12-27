from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from pyxirr import xirr

def run(helpers,drive_service, sheet_service, gc):
    users = []
    response = drive_service.files().list(
        q="name = 'Finance' and mimeType = 'application/vnd.google-apps.folder' and trashed = false",
        fields="nextPageToken, files(id, name)"
    ).execute()
    for finance_folder in response.get('files', []):
        f_id = finance_folder['id']
        zerodha_id = helpers.get_folder_id(drive_service, f_id,'Zerodha')
        personal = helpers.get_file_id(drive_service, f_id, 'Personal Finance')
        
        dividend_id = helpers.get_folder_id(drive_service, zerodha_id, 'Dividend Statement')
        contract_id = helpers.get_folder_id(drive_service, zerodha_id, 'Contract Note')
        
        users.append({
            'finance': f_id,
            'zerodha': zerodha_id,
            'dividend': dividend_id,
            'contract': contract_id,
            'personal': personal
        })

    for user in users:
        account = None

        dividend_statement = [['Symbol','Date','Net Amount','Year-Month']]
        for file in helpers.list_files_in_folder(drive_service, user['dividend']):
            if account is None:
                account = file['name'].split('-')[1]
            file_content = helpers.download_file_content(drive_service, file['id'])
            workbook = load_workbook(file_content)
            worksheet = workbook['Equity Dividends']
            found_header = False
            for row in worksheet.iter_rows(values_only=True):
                if not found_header:
                    if 'Symbol' in row: 
                        found_header = True
                    continue
                if row[1] == 'Total Dividend Amount':
                    break
                dividend_statement.append(
                    [row[1].replace('#','').replace('6',''),
                    row[3], 
                    row[6], 
                    row[3][:7]
                ])
        dividend_statement[1:] = sorted(dividend_statement[1:], key=lambda row: datetime.strptime(row[1], "%Y-%m-%d"), reverse=True)

        contract_note = [['Symbol','Date','Net Amount','Year-Month']]
        for file in helpers.list_files_in_folder(drive_service, user['contract']):
            file_content = helpers.download_file_content(drive_service, file['id'])
            workbook = load_workbook(file_content)
            for worksheet in workbook.worksheets:
                header_found = False
                skip_next_row = False
                sheet_date = datetime.strptime(worksheet.title, "%d-%m-%Y").strftime("%Y-%m-%d")
                for row in worksheet.iter_rows(values_only=True):
                    if not header_found:
                        if row[0] == 'Order No.':
                            header_found = True
                            skip_next_row = True
                        continue
                    if skip_next_row:
                        skip_next_row = False
                        continue
                    if not row[0] or row[0] == 'PAY IN / PAY OUT OBLIGATION':
                        break
                    contract_note.append([
                        row[4].split(' - ')[0], 
                        sheet_date,
                        row[11] if row[5]=='buy' else row[12], 
                        sheet_date[:7]
                    ])
        contract_note[1:] = sorted(contract_note[1:], key=lambda row: datetime.strptime(row[1], "%Y-%m-%d"), reverse=True)

        holdings = [['Symbol','Quantity','Invested','Current']]
        holdings_file_id = helpers.get_file_id(drive_service, user['zerodha'], 'holdings.csv')
        csv_content = helpers.download_file_content(drive_service, holdings_file_id).read().decode('utf-8')
        lines = csv_content.splitlines()
        for line in lines[1:]:
            parts = line.strip().split(',')
            holdings.append([parts[0].replace('"',''), parts[1], parts[4], parts[5]])

        capital_gains = [['Symbol','Date','Invested','Unrealised','Realised','XIRR']]
        symbol_to_contracts = {}
        symbol_to_holdings = {}
        for contract in contract_note[1:]:
            symbol = contract[0]
            if symbol not in symbol_to_contracts: 
                symbol_to_contracts[symbol] = []
            symbol_to_contracts[symbol].append(contract)
        for holding in holdings[1:]:
            symbol_to_holdings[holding[0]] = holding
        for symbol,contracts in symbol_to_contracts.items():
            holdings_invested = 0
            holdings_current = 0
            contracts_invested_held = 0
            contracts_invested_sold = 0
            contracts_sold = 0
            dates = []
            flows = []
            if symbol in symbol_to_holdings:
                holdings_invested = float(symbol_to_holdings[symbol][2])
                holdings_current = float(symbol_to_holdings[symbol][3])
            for c in contracts:
                amt = float(c[2])
                d_obj = datetime.strptime(c[1], "%Y-%m-%d").date()
                if amt < 0:
                    if holdings_invested > contracts_invested_held:
                        contracts_invested_held += amt
                        unrealised = (abs(amt)/holdings_invested)*holdings_current
                        xirr_per = round(xirr([d_obj, datetime.today().date()], [amt, unrealised])*100, 2)
                        capital_gains.append([symbol, c[1], amt, unrealised+amt, 0, xirr_per])
                    else:
                        dates.append(d_obj)
                        flows.append(amt)
                        contracts_invested_sold += amt
                else:
                    dates.append(d_obj)
                    flows.append(amt)
                    contracts_sold += amt
            if contracts_sold > 0:
                realised = contracts_sold + contracts_invested_sold
                xirr_per = round(xirr(dates, flows)*100, 2)
                capital_gains.append([symbol, None, contracts_invested_sold, 0, realised, xirr_per])
        capital_gains.append(['LTCG',datetime.now().replace(year=datetime.now().year-1).strftime("%Y-%m-%d"),None,None,None,None])
        capital_gains[1:] = sorted(capital_gains[1:], key=lambda row: (row[1] is not None, datetime.strptime(row[1], "%Y-%m-%d") if row[1] else None), reverse=True)

        doc = gc.open_by_key(user['personal'])
        sheet_to_data = {
            'Dividend Statement': dividend_statement,
            'Contract Note': contract_note,
            'Holdings': holdings,
            'Capital Gains': capital_gains
        }
        for sheet, data in sheet_to_data.items():
            worksheet = doc.worksheet(sheet)
            worksheet.clear()
            worksheet.update('A1', data)

        doc = gc.open('Screener Tracker')
        stocks = doc.worksheet('Stocks')
        fvs = sheet_service.spreadsheets().get(spreadsheetId=doc.id,fields="sheets(filterViews(filterViewId,title))").execute()
        for s in fvs.get("sheets", []):
            for fv in s.get("filterViews", []):
                if fv.get("title") in ['Growth Stocks',account]:
                    sheet_service.spreadsheets().batchUpdate(
                        spreadsheetId=doc.id,
                        body={"requests": [{"deleteFilterView": {"filterId": fv["filterViewId"]}}]}
                    ).execute()
        resp = sheet_service.spreadsheets().batchUpdate(
            spreadsheetId=doc.id,
            body={"requests": [
            {"addFilterView":
            {"filter":
            { "title": "Growth Stocks",
                "range": {
                    "sheetId": stocks.id,
                    "startRowIndex": 0,
                    "endRowIndex": stocks.row_count,
                    "startColumnIndex": 0,
                    "endColumnIndex": stocks.col_count
                },
                "criteria": {
                    "6": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=G2>1000"}]}},
                    "7": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=H2>1000"}]}},
                    "8": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=I2>10"}]}},
                    "9": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=OR(ISBLANK(J2), J2>10)"}]}},
                    "10": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=OR(ISBLANK(K2), K2>10)"}]}},
                    "11": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=OR(ISBLANK(L2), L2>10)"}]}},
                    "12": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=M2>10"}]}},
                    "13": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=OR(ISBLANK(N2), N2>10)"}]}},
                    "14": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=OR(ISBLANK(O2), O2>10)"}]}},
                    "15": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=OR(ISBLANK(P2), P2>10)"}]}},
                    "16": { "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [{"userEnteredValue": "=Q2>10"}]
            }}}}}},
            {"addFilterView": {
                "filter": {
                    "title": account,
                    "range": {
                        "sheetId": stocks.id,
                        "startRowIndex": 0,
                        "endRowIndex": stocks.row_count,
                        "startColumnIndex": 0,
                        "endColumnIndex": stocks.col_count
                    },
                    "criteria": {
                        "0": {
                            "hiddenValues": list({r[0] for r in sheet_service.spreadsheets().values().get(spreadsheetId=doc.id, range='Stocks!A:A').execute().get('values', []) if r}-set(symbol_to_holdings.keys()))
        }}}}}]}).execute()